from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ExcelHandler:
    def __init__(self, excel_path):
        self.excel_path = excel_path

    def mark_green(self, sheet_name, row_index):
        """
        Mark the specified cell in the Excel sheet with green color.
        """
        wb = load_workbook(self.excel_path)
        sheet = wb[sheet_name]
        light_green_fill = PatternFill(start_color="C6E2A4", end_color="C6E2A4", fill_type="solid")
        column = 5
        cell = sheet.cell(row=row_index + 2, column=column)
        cell.fill = light_green_fill
        wb.save(self.excel_path)
        print(f"Cell {cell} filled with light green color")
